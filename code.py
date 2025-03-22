
from tkinter import *
from tkinter import filedialog
import ttkbootstrap as tb
from ttkbootstrap.scrolled import ScrolledFrame
from ttkbootstrap.toast import ToastNotification
from ttkbootstrap.dialogs import Messagebox
import sqlite3
import csv
import os
from openpyxl import Workbook
import matplotlib.pyplot as plt
from datetime import datetime
from collections import defaultdict


root = tb.Window(title="Expense Tracker",
                 themename="superhero",
                 iconphoto="image/84b81735b361bd85ac4cfef337c88308-removebg-preview.png",
                 size=(600,600))

#classes

class DB_work:

    entries = [] #stores all amount entries 
    combos = [] #store all combo boxes
    dynamic_frames = []  # stores dynamically created frames

    def __init__(self):
        try:
            
            # Get the path to the user's Documents folder
            document_path = os.path.join(os.path.expanduser("~"),"Downloads")

            # Specify the database location in the Documents folder
            db_path = os.path.join(document_path,"exptracker.db")
            conn = sqlite3.connect(db_path)
            
            #connecting DB
            ## Get the path to the user's Documents folder
            document_path = os.path.join(os.path.expanduser("~"),"Downloads")

            # Specify the database location in the Documents folder
            db_path = os.path.join(document_path,"exptracker.db")
            conn = sqlite3.connect(db_path)
            query = conn.cursor()

            #Create tables if they don't exist
            query.execute("""create table IF NOT EXISTS DATA(
                        entry_date Date PRIMARY KEY,
                        salary REAL,
                        additional_income REAL DEFAULT NULL,
                        travel REAL DEFAULT NULL,
                        food REAL DEFAULT NULL,
                        regular_expense REAL DEFAULT NULL
                        )""")
            
            query.execute("""create table IF NOT EXISTS Income_expenditure(
                        ITEMs TEXT PRIMARY KEY,
                        Inc_exp TEXT
                        )""")

            # Insert default items into Income_expenditure
            default_items = [
                ('additional_income', 'income'),
                ('travel', 'expense'),
                ('food', 'expense'),
                ('regular_expense', 'expense')
            ]

            for item in default_items:
                query.execute("INSERT OR IGNORE INTO Income_expenditure (ITEMs, Inc_exp) VALUES (?, ?)", item)

            conn.commit()

        except sqlite3.Error as e:

            toast = ToastNotification(title="Expense tracker warning message ⚠️",
                                      message=f"An error occurred while working with the database: {e}",
                                      duration=3000,
                                      bootstyle="danger",
                                      alert=True,
                                      ) 
            toast.show_toast()

        finally:
            # Ensure the connection is closed, even if an error occur
            if conn:
                conn.close()

    @classmethod
    def new_section_of_label(cls):
        
        #connecting DB
        # Get the path to the user's Documents folder
        document_path = os.path.join(os.path.expanduser("~"),"Downloads")

        # Specify the database location in the Documents folder
        db_path = os.path.join(document_path,"exptracker.db")
        conn = sqlite3.connect(db_path)
        query = conn.cursor()

        query.execute("select ITEMs from Income_expenditure")
        items_col = query.fetchall()

        # Create a new frame inside the same main frame to hold the new row
        new_set_frame = tb.Frame(main_ex_frame)
        new_set_frame.pack(fill="x", pady=5)
        
        # Add 1st entry in the new row frame
        cls.Exp_amount_entry = tb.Entry(new_set_frame,bootstyle="success")
        cls.Exp_amount_entry.pack(pady=10,side="left",padx=5)

        #creating expenditure category combobox

        cls.items = set()
        for i in items_col:
            val = i[0].replace("_"," ").strip()
            cls.items.add(val)

        # Add 2nd entry in the new row frame, on the same line
        cls.category_combobox = tb.Combobox(new_set_frame,bootstyle="success",values=sorted(list(cls.items)))
        cls.category_combobox.pack(side="left",pady=10,padx=5)

        cls.category_combobox.current(0)

        # Save references for later use
        cls.entries.append(cls.Exp_amount_entry)
        cls.combos.append(cls.category_combobox)
        cls.dynamic_frames.append(new_set_frame)  # Add the frame to the dynamic_frames list

        conn.commit()
        conn.close()


    @classmethod
    def submit(cls):    

        # Get the path to the user's Documents folder
        document_path = os.path.join(os.path.expanduser("~"),"Downloads")

            # Specify the database location in the Documents folder
        db_path = os.path.join(document_path,"exptracker.db")
        conn = sqlite3.connect(db_path)
        query =  conn.cursor()

        query.execute("select * from DATA where entry_date = ?",(Date_Entry.entry.get(),))
        ck_data_presented = query.fetchone()

        
        if not ck_data_presented:
            query.execute("insert into DATA (entry_date,salary) values(?,?)",(Date_Entry.entry.get(),salary_Entry.get()))

        # Loop through the dynamically added entries and combo boxes
        for i in range(len(cls.entries)):
            amount = cls.entries[i].get()
            category = cls.combos[i].get()

            print(amount)
            print(category)
            
            if category in cls.items:
                #create a function which need to add record if the record already has a value
                caty_item2 = category.replace(" ", "_").strip().lower() 

                query.execute(f"SELECT {caty_item2} FROM DATA WHERE entry_date = ?", (Date_Entry.entry.get(),))
                ck_1 = query.fetchone()

                # If the category is None, insert the amount
                if ck_1 is None or ck_1[0] is None:
                    query.execute(f"UPDATE DATA SET {caty_item2} = ? WHERE entry_date = ?", (amount, Date_Entry.entry.get()))
                else:
                    existing_value = ck_1[0]

                    # Handle None or empty existing value
                    if not existing_value:
                        existing_value = 0

                    # Safely add the new value
                    try:
                        new_value = float(existing_value) + float(amount)
                        query.execute(f"UPDATE DATA SET {caty_item2} = ? WHERE entry_date = ?", (new_value, Date_Entry.entry.get()))
                    except ValueError:
                        print(f"Invalid data in database or input: {existing_value} or {amount}")

        for i in range(len(cls.entries)):
            amount = cls.entries[i].get()
            category = cls.combos[i].get()

            if category not in cls.items:
                DB_work.category_pop(category,amount)

        conn.commit()
        conn.close()
        
        for frame in cls.dynamic_frames:
            frame.destroy()

        # Clear the lists after destroying the frames
        cls.entries.clear()
        cls.combos.clear()
        cls.dynamic_frames.clear()

        Date_Entry.entry.delete(0,END)
        salary_Entry.delete(0,END)

    @classmethod
    def category_pop(cls,caty,amt):
        
        new_category_pop = tb.Toplevel(title="new category insert",size=(300,300))

        the_category = caty
        amt_of_newcat = amt
        
        new_category_entry = tb.Entry(new_category_pop)
        new_category_entry.pack(pady=20)
        new_category_entry.insert(0,the_category)

        inc_or_exp = StringVar()

        income_radio_btn = tb.Radiobutton(new_category_pop,variable=inc_or_exp,value="income",bootstyle = "success",text="income")
        income_radio_btn.pack(pady=20)

        expense_radio_btn = tb.Radiobutton(new_category_pop,variable=inc_or_exp,value="expense",bootstyle = "danger",text="expense")
        expense_radio_btn.pack(pady=20)

        cp_submit = tb.Button(new_category_pop,bootstyle="success",text="ADD",command=lambda: DB_work.alt_tables(the_category,inc_or_exp.get(),amt_of_newcat,new_category_pop))
        cp_submit.pack(pady=20)

    @classmethod 
    def alt_tables(cls,caty,value,amt,pop_window):

        caty_item = caty.replace(" ","_").strip().lower()

        the_amount = amt
        
        # Get the path to the user's Documents folder
        document_path = os.path.join(os.path.expanduser("~"),"Downloads")

            # Specify the database location in the Documents folder
        db_path = os.path.join(document_path,"exptracker.db")
        conn = sqlite3.connect(db_path)
        query = conn.cursor()

        query.execute("insert into Income_expenditure values (?,?)",(caty_item,value))

        query.execute(f"ALTER TABLE DATA ADD COLUMN {caty_item} REAL DEFAULT NULL")

        query.execute(f"update DATA set {caty_item} = ? where entry_date = ?",(the_amount,Date_Entry.entry.get()))

        conn.commit()
        conn.close()
        
        pop_window.destroy()

#functions

def select_theme(x):
    root.style.theme_use(x)


def get_DB_chart(fdate,tdate):

    # Get the path to the user's Documents folder
    document_path = os.path.join(os.path.expanduser("~"),"Downloads")

    # Specify the database location in the Documents folder
    db_path = os.path.join(document_path,"exptracker.db")
    conn = sqlite3.connect(db_path)
    query = conn.cursor()

    if fdate == tdate:

        query.execute("select * from DATA where entry_date = ?",(fdate,))
        fetch_data = query.fetchall()

        if not fetch_data:
            toaster = ToastNotification(title="Expense Tracker notification",
                                        message=f"There is no data to fetch on {fdate}",
                                        duration=3000,
                                        alert=True,
                                        bootstyle="warning")
            
            toaster.show_toast()
        else:

            #print(fetch_data)
            query.execute("PRAGMA table_info(DATA)")
            column = query.fetchall()
            
            category = [col[1] for col in column if col[1] not in ["entry_date","salary"]]

            query.execute(f"select "+", ".join(category)+" from DATA where entry_date = ?",(fdate,))
            amounts = query.fetchone()
            print(category)
            print(amounts)

            # Filter out None values
            filtered_categories = []
            filtered_amounts = []
            for cat, amt in zip(category, amounts):
                if amt is not None:
                    filtered_categories.append(cat)
                    filtered_amounts.append(amt)

            if amounts:
                plt.bar(filtered_categories, filtered_amounts)
                plt.xlabel("Category")
                plt.ylabel("Amount")
                plt.title(f"Expense Data for {fdate}")
                plt.xticks(rotation=45)
                plt.tight_layout()
                plt.show()
    
    if fdate != tdate:
        # Fetch data between the two dates
        query.execute("""
            SELECT * FROM DATA
            WHERE entry_date BETWEEN ? AND ?
            ORDER BY substr(entry_date, 7, 4) ASC,  -- Year
                     substr(entry_date, 4, 2) ASC,  -- Month
                     substr(entry_date, 1, 2) ASC;  -- Day
        """, (fdate, tdate))
        fetch_data = query.fetchall()

        print(fetch_data)

        if not fetch_data:
            print(f"No data available between {fdate} and {tdate}.")
            return

        # Prepare data for multiline chart
        dates = [datetime.strptime(row[0], "%d-%m-%Y") for row in fetch_data]
        
        # Get column names excluding entry_date and salary
        query.execute("PRAGMA table_info(DATA)")
        categories = [col[1] for col in query.fetchall() if col[1] not in ["entry_date", "salary"]]
        
        category_data = {category: [] for category in categories}

        # Populate category_data with amounts, replace None with 0
        for row in fetch_data:
            for idx, category in enumerate(categories, start=1):
                category_data[category].append(row[idx] if row[idx] is not None else 0)

        # Multiline chart for each category over dates
        plt.figure(figsize=(10, 6))
        for category, amounts in category_data.items():
            plt.plot(dates, amounts, marker='o', label=category)
        plt.xlabel("Date")
        plt.ylabel("Amount")
        plt.title(f"Expense Trends from {fdate} to {tdate}")
        plt.xticks(rotation=45)
        plt.legend(title="Categories")
        plt.tight_layout()
        plt.show()

        # Pie chart for total amounts per category in date range
        total_amounts = [sum(amounts) for amounts in category_data.values()]
        filtered_totals = [amt for amt in total_amounts if amt > 0]  # Exclude categories with zero sum
        filtered_categories = [category for category, amt in zip(categories, total_amounts) if amt > 0]

        if filtered_totals:
            plt.figure(figsize=(8, 8))
            plt.pie(filtered_totals, labels=filtered_categories, autopct='%1.1f%%', startangle=140)
            plt.title(f"Category Breakdown from {fdate} to {tdate}")
            plt.show()

    conn.commit()
    conn.close()

def upload_csv():
    try:
        CSV_frame.filename = filedialog.askopenfilename(initialdir="C:/users/", title="CSV files", filetypes=[("csv files", "*csv")])
        if CSV_frame.filename:  # Check if a file was selected
            # Dictionary to hold data grouped by expense category
            expense_data = defaultdict(list)
            
            # Open the CSV file and read data
            with open(CSV_frame.filename, newline='', encoding='utf-8-sig') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    # Parse the date (format: MM/DD/YY)
                    parsed_date = datetime.strptime(row['Date'], "%m/%d/%y")
                    amount = float(row['Debit'])  # Convert 'Debit' to float
                    expense = row['Expense']  # Category of expense
                    expense_data[expense].append((parsed_date, amount))

            # Aggregate total expenses by category for the pie chart
            category_totals = {expense: sum(amount for _, amount in data) for expense, data in expense_data.items()}

            # Create subplots: 1 row, 2 columns
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6))

            # Multiline Plot (Left)
            for expense, data in expense_data.items():
                # Sort data by date for each category
                data.sort()
                dates, amounts = zip(*data)  # Unpack dates and amounts
                ax1.plot(dates, amounts, marker='o', label=expense)  # Plot each category

            # Format multiline plot
            ax1.set_xlabel("Date")
            ax1.set_ylabel("Amount")
            ax1.set_title("Expenses over Time by Category")
            ax1.tick_params(axis='x', rotation=45)
            ax1.legend(title="Expense Category", bbox_to_anchor=(1.05, 1), loc='upper left')  # Position legend outside plot

            # Pie Chart (Right)
            ax2.pie(category_totals.values(), labels=category_totals.keys(), autopct='%1.1f%%', startangle=140)
            ax2.set_title("Total Expense Distribution by Category")

            # Adjust layout to prevent overlap
            plt.tight_layout()
            plt.show()

    except Exception  as e :

        Messagebox.show_warning(title="Expense Tracker",message="There is a problem with in the file",alert=True)
            


def download_xl():

    try:

        current_time = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")

        #print(current_time)

        default_path = os.path.join(os.path.expanduser("~"),"Downloads")

        file_name = f"EXPENSE_Tracker_{current_time}.xlsx"

        file_path = os.path.join(default_path,file_name)

        wb = Workbook()

        ws1 =  wb.create_sheet(title="DATA",index=0)

        ws2 = wb.create_sheet(title="Income expenditure",index=1)

        # Get the path to the user's Documents folder
        document_path = os.path.join(os.path.expanduser("~"),"Downloads")

        # Specify the database location in the Documents folder
        db_path = os.path.join(document_path,"exptracker.db")
        conn = sqlite3.connect(db_path)
        query = conn.cursor()
        
        query.execute("PRAGMA table_info(DATA)")
        data_table_col = query.fetchall()

        ws1.append([dtr[1] for dtr in data_table_col])

        query.execute("""
            SELECT * FROM DATA
            ORDER BY substr(entry_date, 7, 4) ASC,  -- Year
                    substr(entry_date, 4, 2) ASC,  -- Month
                    substr(entry_date, 1, 2) ASC;  -- Day
        """)
        data = query.fetchall()

        for row in data:
            ws1.append(list(row))

        query.execute("PRAGMA table_info(Income_expenditure)")
        IE_table_col = query.fetchall()

        ws2.append([row[1] for row in IE_table_col])    
            

        query.execute("select * from Income_expenditure order by ITEMs")
        ie_data = query.fetchall()
        
        for row in ie_data:
            ws2.append(list(row))

        conn.commit()

        Success_toast = ToastNotification(title="Expense tracker notification",
                                      message=f"the excel file successfully downloaded in documents",
                                      duration=3000,
                                      bootstyle="success",
                                      alert=True,
                                      ) 
        Success_toast.show_toast()
    except Exception as e:
        toast = ToastNotification(title="Expense tracker warning message ⚠️",
                                      message=f"An error occurred while downloading excel file: {e}",
                                      duration=3000,
                                      bootstyle="warning",
                                      alert=True,
                                      ) 
        toast.show_toast()
        
    finally:
        if conn:
            conn.close()

    wb.save(file_path)


def enable_disable_date(get_date):

    selected_date = get_date.get()

    if selected_date == "":
        return

    selected_date_obj = datetime.strptime(selected_date,"%d-%m-%Y")
    
    if selected_date_obj.day == 1:
        salary_Entry.config(state="normal") 
    else:
        salary_Entry.config(state="disabled")

def dynamic_update_DB(get_date):
    selected_date = get_date.get()  # Get the date from the user interface
    
    # Check if the date is empty
    if not selected_date.strip():
        print("No date selected")
        return

    try:
        selected_date_obj = datetime.strptime(selected_date, "%d-%m-%Y")
    except ValueError:
        print("Invalid date format. Please provide a date in DD-MM-YYYY format.")
        return

    document_path = os.path.join(os.path.expanduser("~"), "Downloads")
    db_path = os.path.join(document_path, "exptracker.db")
    conn = sqlite3.connect(db_path)
    query = conn.cursor()

    try:
        # Fetch entries for the selected month and year
        query.execute("""
            SELECT * FROM DATA 
            WHERE substr(entry_date, 4, 2) || '-' || substr(entry_date, 7, 4) = ? 
            ORDER BY entry_date
        """, (selected_date_obj.strftime('%m-%Y'),))
        entries = query.fetchall()

        # Get column names from the DATA table
        query.execute("PRAGMA table_info(DATA)")
        columns = query.fetchall()
        column_names = [col[1] for col in columns]

        query.execute("SELECT ITEMs, Inc_exp FROM Income_expenditure")
        inc_exp_data = query.fetchall()

        # Separate income and expense columns
        income_columns = [item[0] for item in inc_exp_data if item[1] == 'income']
        expense_columns = [item[0] for item in inc_exp_data if item[1] == 'expense']

        if entries:
            for i, entry in enumerate(entries):
                e_date = entry[0]
                d_salary = entry[1] if entry[1] is not None else 0.0

                total_income = sum([float(entry[column_names.index(col)] or 0) for col in income_columns if col in column_names]) + float(d_salary or 0)
                total_expense = sum([float(entry[column_names.index(col)] or 0) for col in expense_columns if col in column_names])


                net_income = total_income - total_expense
                print(f"Entry Date: {e_date}, Salary: {d_salary}, Total Income: {total_income}, Total Expense: {total_expense}, Net Income: {net_income}")

                # Update the salary of the next entry within the same month
                if i + 1 < len(entries):
                    next_entry_date = entries[i + 1][0]
                    query.execute("""
                        UPDATE DATA
                        SET salary = ?
                        WHERE entry_date = ?
                    """, (net_income, next_entry_date))
                    print(f"Updated next entry's salary for date {next_entry_date} to {net_income}")

    except sqlite3.Error as e:
        print(f"Database error: {e}")
    finally:
        conn.commit()  # Ensure all changes are saved
        conn.close()


#GUI Title

title_label = tb.Label(root,text="Expense Tracker",font=("Poor Richard",38))
title_label.pack(pady=20)

#menubutton label frame
theme_menu_label_frame =tb.Frame(root)
theme_menu_label_frame.pack(pady=10)

#theme menubutton label
menu_label = tb.Label(theme_menu_label_frame,text="Select the theme:")
menu_label.pack(padx=5,side="left")

#theme menubutton
theme_menu = tb.Menubutton(theme_menu_label_frame,bootstyle="info",text="theme")
theme_menu.pack(padx=5,side="left")

#create a menu
select_menu = tb.Menu(theme_menu)

theme_var = StringVar()
for theme in ["cosmo","flatly","litera","minty","lumen","yeti","pulse","united","morph","journal","darkly","superhero","solar","cyborg","vapor","simplex","cerculean"]:
    select_menu.add_radiobutton(label=theme,variable=theme_var,command= lambda x = theme: select_theme(x))

theme_menu["menu"] = select_menu

#creating a notebook

note_book = tb.Notebook(root,bootstyle="info")
note_book.pack(pady=40,padx=20,fill="both",expand=1)

#creating frames
DB_frame = tb.Frame(note_book)
CSV_frame = tb.Frame(note_book)
XL_frame = tb.Frame(note_book)

#****************************************************************************************************************************************#
#DB_Frame Section

#creating scrolled frame

DB_Scrolled_frame = ScrolledFrame(DB_frame)
DB_Scrolled_frame.pack(fill=BOTH,expand=True)

intk = DB_work()

#defualt labels and entry

Date_label = tb.Label(DB_Scrolled_frame,text="Date:",font=('Helvertica',18))
Date_label.pack(pady=10)

Date_Entry = tb.DateEntry(DB_Scrolled_frame,firstweekday=6,bootstyle="info")
Date_Entry.pack()

date_get = StringVar()
date_get.trace_add("write",lambda name, mode, value, get_date = date_get: enable_disable_date(get_date))
date_get.trace_add("write",lambda name, mode, value, get_date = date_get: dynamic_update_DB(get_date))
Date_Entry.entry.config(textvariable=date_get)

Salary_label = tb.Label(DB_Scrolled_frame,text="Salary:",font=('Helvertica',18))
Salary_label.pack(pady=10)

salary_Entry = tb.Entry(DB_Scrolled_frame,state="disabled")
salary_Entry.pack(pady=10,ipadx=16)

#income and expenditure labels,Entry,combobutton

#creating expenditure category label
category_label = tb.Label(DB_Scrolled_frame,text= "Enter the amount and set the category:",font=("Helvertica",18))
category_label.pack(pady=20)

#creating a main frame to hold multiple frames 
main_ex_frame = tb.Frame(DB_Scrolled_frame)
main_ex_frame.pack(pady=5)

#1st frame in main frame and setting a frame to bring the amount entry and category menubutton on a same line
set_frame = tb.Frame(main_ex_frame)
set_frame.pack(fill="x", pady=5)


#add button to insert multiple income and expenditure labels,Entry,combobox

exp_add_button = tb.Button(DB_Scrolled_frame,text="+",bootstyle = "info",command=DB_work.new_section_of_label)
exp_add_button.pack()

#submit buttons

submit_button = tb.Button(DB_Scrolled_frame,text="Submit",bootstyle="success outline",command=DB_work.submit)
submit_button.pack(pady=20)

#separator

my_sep = tb.Separator(DB_Scrolled_frame,bootstyle="light",orient="horizontal")
my_sep.pack(fill="x",padx=40)

#Data visualization section
DV_title_label = tb.Label(DB_Scrolled_frame,text="Visualize your Expense",font=("Helvertica",22))
DV_title_label.pack(pady=20)

From_date = tb.Label(DB_Scrolled_frame,text="From :",font=("Helvertica", 18))
From_date.pack(pady=10)

From_date_entry = tb.DateEntry(DB_Scrolled_frame,firstweekday=6,bootstyle="warning")
From_date_entry.pack(pady=10)


to_date = tb.Label(DB_Scrolled_frame,text="To :",font=("Helvertica", 18))
to_date.pack(pady=10)

to_date_entry = tb.DateEntry(DB_Scrolled_frame,firstweekday=6,bootstyle="warning")
to_date_entry.pack(pady=10)

DV_submit_btn = tb.Button(DB_Scrolled_frame,text="View Data",bootstyle = "success",command= lambda:get_DB_chart(From_date_entry.entry.get(),to_date_entry.entry.get()))
DV_submit_btn.pack(pady=20)

#******************************************************************************************************************************#

#CSV_Frame Section

upload_csv_btn = tb.Button(CSV_frame,bootstyle="reverse light",text="Upload CSV",command=upload_csv)
upload_csv_btn.pack(pady=20)

#****************************************************************************************************************************#
#XL_Frame Section

export_xl_btn = tb.Button(XL_frame,bootstyle="success outline",text="Export Excel",command=download_xl)
export_xl_btn.pack(pady=20)

#*****************************************************************************************************************************#
#add frame to notebook
note_book.add(DB_frame,text="Database")
note_book.add(CSV_frame,text="Upload Csv")
note_book.add(XL_frame,text="Export Excel")

root.mainloop()